import os

# Monkey-patch eventlet before other imports when running on Render
if os.environ.get('RENDER'):
    import eventlet
    eventlet.monkey_patch()

import json
from datetime import datetime
from io import BytesIO
import zipfile
import tempfile
import requests

from flask import Flask, render_template, jsonify, request, send_file
from werkzeug.utils import secure_filename
from flask_socketio import SocketIO, emit
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')

async_mode = 'eventlet' if os.environ.get('RENDER') else 'threading'
socketio = SocketIO(app, cors_allowed_origins="*", async_mode=async_mode)

# Use Render disk for persistence on Render, local data dir otherwise
if os.environ.get('RENDER'):
    DATA_DIR = '/var/data'
else:
    DATA_DIR = os.path.join(os.path.dirname(__file__), 'data')

# Ensure data directory exists
os.makedirs(DATA_DIR, exist_ok=True)
print(f"Data directory: {DATA_DIR}, exists: {os.path.exists(DATA_DIR)}")

DATA_FILE = os.path.join(DATA_DIR, 'schedule.json')
HISTORY_FILE = os.path.join(DATA_DIR, 'course_history.json')

# Valid terms
VALID_TERMS = ['fall-2026', 'spring-2027']
DEFAULT_TERM = 'fall-2026'

# Reset password - set via PASSWORD or password environment variable
RESET_PASSWORD = os.environ.get('PASSWORD', '') or os.environ.get('password', '')

# Course guides data - courses students should take together each semester
COURSE_GUIDES = {
    'business': {
        'name': 'Business Administration',
        'semesters': {
            'Y1-Fall': ['MGMT 205', 'ECON 205'],
            'Y1-Spring': ['ECON 206'],
            'Y2-Fall': ['MGMT 281', 'ECMG 303', 'MGMT 301'],
            'Y2-Spring': ['MGMT 284', 'MGMT 306', 'ECON 306'],
            'Y3-Fall': ['MGMT 312'],
            'Y3-Spring': ['MGMT 314', 'MGMT 399'],
            'Y4-Fall': ['ITMG 388', 'MGMT 454'],
            'Y4-Spring': ['MGMT 411']
        }
    },
    'accounting': {
        'name': 'Accounting',
        'semesters': {
            'Y1-Fall': ['MGMT 205', 'ECON 205', 'MGMT 281'],
            'Y1-Spring': ['MGMT 284', 'ECON 206'],
            'Y2-Fall': ['MGMT 321', 'MGMT 306'],
            'Y2-Spring': ['MGMT 322', 'MGMT 312'],
            'Y3-Fall': ['MGMT 432', 'MGMT 314', 'ECMG 303'],
            'Y3-Spring': ['MGMT 331', 'MGMT 433', 'ITMG 388'],
            'Y4-Fall': ['MGMT 434', 'MGMT 399', 'MGMT 454'],
            'Y4-Spring': ['MGMT 411']
        }
    },
    'economics': {
        'name': 'Economics',
        'semesters': {
            'Y1-Fall': ['ECON 205'],
            'Y1-Spring': ['ECON 206'],
            'Y2-Fall': ['ECON 305'],
            'Y2-Spring': ['ECON 306'],
            'Y3-Fall': ['ECON 452'],
            'Y4-Fall': ['ECON 480'],
            'Y4-Spring': ['ECON 470']
        }
    },
    'finance': {
        'name': 'Finance',
        'semesters': {
            'Y1-Fall': ['MGMT 205', 'ECON 205'],
            'Y1-Spring': ['ECON 206'],
            'Y2-Fall': ['MGMT 281', 'ECMG 303', 'MGMT 301'],
            'Y2-Spring': ['MGMT 284', 'MGMT 306', 'MGMT 312'],
            'Y3-Fall': ['MGMT 314', 'MGMT 402', 'MGMT 454'],
            'Y3-Spring': ['MGMT 370', 'ITMG 388', 'MGMT 410'],
            'Y4-Fall': ['MGMT 411'],
            'Y4-Spring': ['MGMT 399', 'ECMG 478']
        }
    }
}


def get_schedule_file(term):
    """Get schedule file path for a specific term."""
    if term not in VALID_TERMS:
        term = DEFAULT_TERM
    if term == 'fall-2026':
        return DATA_FILE  # Use existing schedule.json for fall-2026
    return os.path.join(DATA_DIR, f'schedule_{term.replace("-", "_")}.json')


def load_course_history():
    """Load course history data from JSON file."""
    try:
        with open(HISTORY_FILE, 'r') as f:
            return json.load(f)
    except FileNotFoundError:
        return {}


def load_schedule(term=None):
    """Load schedule data from JSON file for a specific term."""
    if term is None:
        term = DEFAULT_TERM
    schedule_file = get_schedule_file(term)
    try:
        with open(schedule_file, 'r') as f:
            data = json.load(f)
            data['term'] = term
            return data
    except FileNotFoundError:
        # For new terms, start with empty schedule but copy faculty/timeSlots from default
        default_data = load_schedule(DEFAULT_TERM) if term != DEFAULT_TERM else {}
        return {
            "term": term,
            "courses": [],
            "instructors": default_data.get("instructors", []),
            "faculty": default_data.get("faculty", []),
            "timeSlots": default_data.get("timeSlots", {})
        }


def save_schedule(data, term=None):
    """Save schedule data to JSON file for a specific term."""
    if term is None:
        term = data.get('term', DEFAULT_TERM)
    schedule_file = get_schedule_file(term)
    os.makedirs(os.path.dirname(schedule_file), exist_ok=True)
    with open(schedule_file, 'w') as f:
        json.dump(data, f, indent=2)


@app.route('/')
def index():
    """Serve the main schedule page."""
    return render_template('index.html')


@app.route('/api/schedule')
def get_schedule():
    """Return current schedule data."""
    term = request.args.get('term', DEFAULT_TERM)
    return jsonify(load_schedule(term))


@app.route('/api/schedule', methods=['POST'])
def update_schedule():
    """Update schedule and broadcast to all clients."""
    data = request.json
    term = data.get('term', DEFAULT_TERM)
    schedule = load_schedule(term)

    course_id = data.get('courseId')
    new_slot_id = data.get('slotId')

    for course in schedule['courses']:
        if course['id'] == course_id:
            course['slotId'] = new_slot_id
            break

    save_schedule(schedule, term)

    socketio.emit('schedule_update', {
        'courseId': course_id,
        'slotId': new_slot_id,
        'term': term,
        'timestamp': datetime.now().isoformat()
    })

    return jsonify({'success': True})


@app.route('/api/course', methods=['POST'])
def update_course():
    """Update course details."""
    data = request.json
    term = data.get('term', DEFAULT_TERM)
    schedule = load_schedule(term)

    course_id = data.get('courseId')
    updates = data.get('updates', {})

    for course in schedule['courses']:
        if course['id'] == course_id:
            for key, value in updates.items():
                if key in course:
                    course[key] = value
            break

    save_schedule(schedule, term)

    socketio.emit('course_update', {
        'courseId': course_id,
        'updates': updates,
        'term': term,
        'timestamp': datetime.now().isoformat()
    })

    return jsonify({'success': True})


@app.route('/api/course/add', methods=['POST'])
def add_course():
    """Add a new course."""
    try:
        data = request.json
        if not data:
            return jsonify({'success': False, 'error': 'No data provided'}), 400

        term = data.get('term', DEFAULT_TERM)
        schedule = load_schedule(term)

        code = data.get('code', '').upper()
        number = str(data.get('number', ''))  # Ensure it's a string
        name = data.get('name', '')
        instructor = data.get('instructor', '')
        room = data.get('room', '')
        slot_id = data.get('slotId')

        if not code or not number:
            return jsonify({'success': False, 'error': 'Code and number are required'}), 400

        # Find the next available section number for this course
        existing_sections = [
            int(c['section']) for c in schedule['courses']
            if c['code'] == code and c['number'] == number and c['section'].isdigit()
        ]
        next_section = str(max(existing_sections, default=0) + 1)

        course_id = f"{code}-{number}-{next_section}"

        new_course = {
            'id': course_id,
            'code': code,
            'number': number,
            'section': next_section,
            'name': name,
            'days': '',
            'startTime': '',
            'endTime': '',
            'instructor': instructor,
            'room': room,
            'slotId': slot_id
        }

        schedule['courses'].append(new_course)
        save_schedule(schedule, term)

        socketio.emit('course_added', {
            'course': new_course,
            'term': term,
            'timestamp': datetime.now().isoformat()
        })

        return jsonify({'success': True, 'course': new_course})
    except Exception as e:
        import traceback
        print(f"Error adding course: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/faculty', methods=['GET'])
def get_faculty():
    """Get all faculty members."""
    term = request.args.get('term', DEFAULT_TERM)
    schedule = load_schedule(term)
    return jsonify(schedule.get('faculty', []))


@app.route('/api/course-history')
def get_course_history():
    """Return all course history data."""
    return jsonify(load_course_history())


@app.route('/api/course-guides')
def get_course_guides():
    """Return course guide data for all programs."""
    return jsonify(COURSE_GUIDES)


@app.route('/api/check-conflicts', methods=['POST'])
def check_conflicts():
    """Check for scheduling conflicts across all programs by year level."""
    data = request.json
    term = data.get('term', DEFAULT_TERM)

    schedule = load_schedule(term)

    # Determine if Fall or Spring
    is_fall = 'fall' in term.lower()
    term_suffix = 'Fall' if is_fall else 'Spring'

    # Build a map of course code to all sections and their slots
    course_sections = {}
    for course in schedule['courses']:
        code_key = f"{course['code']} {course['number']}"
        if code_key not in course_sections:
            course_sections[code_key] = []
        if course.get('slotId'):
            course_sections[code_key].append({
                'section': course['section'],
                'slot': course['slotId'],
                'instructor': course.get('instructor', 'TBA'),
                'room': course.get('room', '')
            })

    # Collect all courses for each year level across ALL programs
    year_levels = ['Y1', 'Y2', 'Y3', 'Y4']
    results_by_year = {}

    for year in year_levels:
        semester_key = f"{year}-{term_suffix}"

        # Gather all unique courses from all programs for this year/semester
        all_courses_for_year = set()
        for program_key, program_data in COURSE_GUIDES.items():
            if semester_key in program_data['semesters']:
                for course in program_data['semesters'][semester_key]:
                    all_courses_for_year.add(course)

        courses_list = list(all_courses_for_year)

        if len(courses_list) == 0:
            continue

        # Check status of each course
        course_status = {}
        for course_code in courses_list:
            sections = course_sections.get(course_code, [])
            if not sections:
                course_status[course_code] = {
                    'scheduled': False,
                    'sections': [],
                    'status': 'not-scheduled'
                }
            else:
                course_status[course_code] = {
                    'scheduled': True,
                    'sections': sections,
                    'status': 'scheduled'
                }

        # Check for conflicts between pairs of courses
        conflicts = []
        for i, course1 in enumerate(courses_list):
            for course2 in courses_list[i+1:]:
                sections1 = course_sections.get(course1, [])
                sections2 = course_sections.get(course2, [])

                if not sections1 or not sections2:
                    continue

                # Check if ALL combinations conflict (ignoring online/asynch sections)
                all_conflict = True
                has_any_conflict = False

                for s1 in sections1:
                    for s2 in sections2:
                        # Skip online/asynch sections - they can overlap
                        room1 = s1.get('room', '').upper()
                        room2 = s2.get('room', '').upper()
                        slot1 = s1.get('slot', '').upper()
                        slot2 = s2.get('slot', '').upper()

                        is_online1 = 'ONLINE' in room1 or 'ASYNCH' in slot1 or slot1 == 'ASYNCH'
                        is_online2 = 'ONLINE' in room2 or 'ASYNCH' in slot2 or slot2 == 'ASYNCH'

                        # If either is online, no physical conflict
                        if is_online1 or is_online2:
                            all_conflict = False
                            continue

                        if s1['slot'] == s2['slot']:
                            has_any_conflict = True
                        else:
                            all_conflict = False

                if has_any_conflict and all_conflict:
                    conflicts.append({
                        'type': 'critical',
                        'courses': [course1, course2],
                        'message': f'{course1} & {course2} - ALL sections conflict'
                    })
                    if course1 in course_status:
                        course_status[course1]['status'] = 'conflict'
                    if course2 in course_status:
                        course_status[course2]['status'] = 'conflict'

        results_by_year[year] = {
            'courses': courses_list,
            'courseStatus': course_status,
            'conflicts': conflicts,
            'label': f"Year {year[1]} {term_suffix}"
        }

    return jsonify({
        'success': True,
        'resultsByYear': results_by_year,
        'term': term,
        'termType': term_suffix
    })


@app.route('/api/course-history/<course_id>')
def get_course_history_by_id(course_id):
    """Return history for a specific course."""
    history = load_course_history()
    if course_id in history:
        return jsonify(history[course_id])
    return jsonify({'error': 'Course not found'}), 404


@app.route('/api/faculty/add', methods=['POST'])
def add_faculty():
    """Add a new faculty member."""
    data = request.json
    term = data.get('term', DEFAULT_TERM)
    schedule = load_schedule(term)

    name = data.get('name', '').strip()
    if not name:
        return jsonify({'success': False, 'error': 'Name is required'}), 400

    if 'faculty' not in schedule:
        schedule['faculty'] = []

    if name not in schedule['faculty']:
        schedule['faculty'].append(name)
        schedule['faculty'].sort()
        save_schedule(schedule, term)

        socketio.emit('faculty_added', {
            'name': name,
            'term': term,
            'timestamp': datetime.now().isoformat()
        })

    return jsonify({'success': True, 'name': name})


@app.route('/api/faculty/delete', methods=['POST'])
def delete_faculty():
    """Delete a faculty member."""
    data = request.json
    term = data.get('term', DEFAULT_TERM)
    schedule = load_schedule(term)

    name = data.get('name', '')
    if 'faculty' in schedule and name in schedule['faculty']:
        schedule['faculty'].remove(name)
        save_schedule(schedule, term)

        socketio.emit('faculty_deleted', {
            'name': name,
            'term': term,
            'timestamp': datetime.now().isoformat()
        })

    return jsonify({'success': True})


@app.route('/api/undo', methods=['POST'])
def undo_action():
    """Undo a previous action."""
    data = request.json
    term = data.get('term', DEFAULT_TERM)
    schedule = load_schedule(term)
    action_type = data.get('type')

    if action_type == 'move':
        course_id = data.get('courseId')
        slot_id = data.get('slotId')
        room = data.get('room')

        for course in schedule['courses']:
            if course['id'] == course_id:
                course['slotId'] = slot_id
                if room is not None:
                    course['room'] = room
                break

        save_schedule(schedule, term)

        socketio.emit('schedule_update', {
            'courseId': course_id,
            'slotId': slot_id,
            'term': term,
            'timestamp': datetime.now().isoformat()
        })

    elif action_type == 'update':
        course_id = data.get('courseId')
        updates = data.get('updates', {})

        for course in schedule['courses']:
            if course['id'] == course_id:
                for key, value in updates.items():
                    course[key] = value
                break

        save_schedule(schedule, term)

        socketio.emit('course_update', {
            'courseId': course_id,
            'updates': updates,
            'term': term,
            'timestamp': datetime.now().isoformat()
        })

    elif action_type == 'delete':
        course_id = data.get('courseId')
        schedule['courses'] = [c for c in schedule['courses'] if c['id'] != course_id]
        save_schedule(schedule, term)

        socketio.emit('course_deleted', {
            'courseId': course_id,
            'term': term,
            'timestamp': datetime.now().isoformat()
        })

    return jsonify({'success': True})


@app.route('/api/export/json')
def export_json():
    """Download schedule as JSON file."""
    term = request.args.get('term', DEFAULT_TERM)
    if term not in VALID_TERMS:
        term = DEFAULT_TERM
    schedule = load_schedule(term)
    buffer = BytesIO()
    buffer.write(json.dumps(schedule, indent=2).encode('utf-8'))
    buffer.seek(0)

    return send_file(
        buffer,
        mimetype='application/json',
        as_attachment=True,
        download_name=f'schedule_{term}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'
    )


@app.route('/api/export/excel')
def export_excel():
    """Download schedule as Excel file in registrar format."""
    term = request.args.get('term', DEFAULT_TERM)
    if term not in VALID_TERMS:
        term = DEFAULT_TERM
    schedule = load_schedule(term)

    # Semester dates
    if term == 'fall-2026':
        start_date = "8/17/2026"
        end_date = "12/8/2026"
    else:
        start_date = "1/11/2027"
        end_date = "5/4/2027"

    # Slot to days/times mapping
    slot_info = {
        'MW-A': {'days': 'MW', 'start': '8:15 AM', 'end': '9:40 AM'},
        'MW-B': {'days': 'MW', 'start': '9:50 AM', 'end': '11:15 AM'},
        'MW-C': {'days': 'MW', 'start': '11:30 AM', 'end': '12:55 PM'},
        'MW-D': {'days': 'MW', 'start': '1:05 PM', 'end': '2:30 PM'},
        'MW-E': {'days': 'MW', 'start': '2:40 PM', 'end': '4:05 PM'},
        'TR-G': {'days': 'TR', 'start': '8:15 AM', 'end': '9:40 AM'},
        'TR-H': {'days': 'TR', 'start': '9:50 AM', 'end': '11:15 AM'},
        'TR-I': {'days': 'TR', 'start': '11:25 AM', 'end': '12:50 PM'},
        'TR-J': {'days': 'TR', 'start': '2:00 PM', 'end': '3:25 PM'},
        'TR-K': {'days': 'TR', 'start': '3:35 PM', 'end': '5:00 PM'},
        'M-EVE': {'days': 'M', 'start': '6:15 PM', 'end': '9:00 PM'},
        'T-EVE': {'days': 'T', 'start': '6:15 PM', 'end': '9:00 PM'},
        'W-EVE': {'days': 'W', 'start': '6:15 PM', 'end': '9:00 PM'},
        'R-EVE': {'days': 'R', 'start': '6:15 PM', 'end': '9:00 PM'},
        'TR-EVE': {'days': 'TR', 'start': '6:15 PM', 'end': '9:00 PM'},
        'SAT': {'days': 'SAT', 'start': '9:00 AM', 'end': '12:00 PM'},
        'ASYNCH': {'days': 'ONLINE', 'start': '12:01 AM', 'end': '12:02 AM'},
    }

    # Core designations (can be expanded)
    core_designations = {
        'ECON 205': 'SocBeh Sci',
        'MGMT 205': 'SocBeh Sci',
        'MGMT 314': 'GP',
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"

    # Headers
    headers = ['Dept.', 'Course', 'Core', 'Section', 'Credit Type', 'Type', 'Course Type',
               'Course Title', 'Scheduled Days', 'Credits', 'Start Time', 'End Time',
               'Start Date', 'End Date', 'Limit', 'Instructor', 'Building', 'Room']

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # Sort courses by code and number
    sorted_courses = sorted(
        [c for c in schedule['courses'] if c.get('slotId')],
        key=lambda x: (x['code'], x['number'], x['section'])
    )

    row = 2
    for course in sorted_courses:
        slot_id = course.get('slotId', '')
        room_full = course.get('room', '')
        course_num = int(course['number']) if course['number'].isdigit() else 0

        # Parse room into building and room number
        if 'ONLINE' in room_full.upper():
            building = 'ONLINE'
            room_num = 'SYNCHR' if 'SYNCHR' in room_full.upper() else 'ASYNCH'
        elif room_full:
            parts = room_full.split()
            building = parts[0] if parts else ''
            room_num = parts[1] if len(parts) > 1 else ''
        else:
            building = ''
            room_num = ''

        # Determine course type
        if slot_id == 'ASYNCH' or 'ASYNCH' in room_full.upper():
            course_type = 'Asynchronous Online'
            building = 'ONLINE'
            room_num = 'ASYNCH'
        elif 'ONLINE' in room_full.upper() and 'SYNCHR' in room_full.upper():
            course_type = 'Synchronous Online'
            building = 'ONLINE'
            room_num = 'SYNCHR'
        else:
            course_type = 'Course'

        # Get slot info for days/times
        info = slot_info.get(slot_id, {})
        days = info.get('days', course.get('days', ''))
        start_time = info.get('start', course.get('startTime', ''))
        end_time = info.get('end', course.get('endTime', ''))

        # Override days for online courses
        if course_type == 'Asynchronous Online':
            days = 'ONLINE'
        elif course_type == 'Synchronous Online' and slot_id:
            # Keep the day letter for synchronous online (e.g., 'R' for Thursday evening)
            if slot_id == 'M-EVE':
                days = 'M'
            elif slot_id == 'T-EVE':
                days = 'T'
            elif slot_id == 'W-EVE':
                days = 'W'
            elif slot_id == 'R-EVE':
                days = 'R'

        # Credit type and class limit based on course level
        if course_num >= 500:
            credit_type = 'CRG'
            core = 'GRMBA'
            limit = 30
        else:
            credit_type = 'CR'
            course_key = f"{course['code']} {course['number']}"
            core = core_designations.get(course_key, '')
            limit = 25 if course_num < 300 else 30

        # Section with leading zero
        section = course['section'].zfill(2)

        # Course code for display
        course_code = f"{course['code']} {course['number']}"

        # Row data
        row_data = [
            'MG',                           # Dept.
            course_code,                    # Course
            core,                           # Core
            section,                        # Section
            credit_type,                    # Credit Type
            'LEC',                          # Type
            course_type,                    # Course Type
            course.get('name', ''),         # Course Title
            days,                           # Scheduled Days
            3,                              # Credits
            start_time,                     # Start Time
            end_time,                       # End Time
            start_date,                     # Start Date
            end_date,                       # End Date
            limit,                          # Limit
            course.get('instructor', ''),   # Instructor
            building,                       # Building
            room_num,                       # Room
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left' if col in [2, 7, 8, 16] else 'center')

        row += 1

    # Auto-size columns
    column_widths = [6, 12, 10, 8, 10, 6, 18, 40, 12, 8, 10, 10, 10, 10, 6, 18, 8, 8]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'schedule_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )


@app.route('/api/backup')
def backup_all_data():
    """Download all data files as a zip archive."""
    buffer = BytesIO()

    with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        # Add all schedule files
        for term in VALID_TERMS:
            schedule_file = get_schedule_file(term)
            if os.path.exists(schedule_file):
                filename = os.path.basename(schedule_file)
                with open(schedule_file, 'r') as f:
                    zf.writestr(filename, f.read())

        # Add course history
        if os.path.exists(HISTORY_FILE):
            with open(HISTORY_FILE, 'r') as f:
                zf.writestr('course_history.json', f.read())

    buffer.seek(0)

    return send_file(
        buffer,
        mimetype='application/zip',
        as_attachment=True,
        download_name=f'scheduler_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.zip'
    )


@app.route('/api/restore', methods=['POST'])
def restore_data():
    """Restore data from uploaded zip file. Requires password."""
    password = request.form.get('password', '').strip()
    expected = RESET_PASSWORD.strip() if RESET_PASSWORD else ''

    # Debug: log password check (don't log actual passwords in production)
    print(f"Password check - entered length: {len(password)}, expected length: {len(expected)}, match: {password == expected}")

    if not expected:
        return jsonify({'success': False, 'error': 'PASSWORD environment variable not set on server'}), 500

    if password != expected:
        return jsonify({'success': False, 'error': 'Invalid password'}), 401

    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'}), 400

    if not file.filename.endswith('.zip'):
        return jsonify({'success': False, 'error': 'File must be a .zip archive'}), 400

    try:
        # Save to temp file and extract
        with tempfile.NamedTemporaryFile(delete=False, suffix='.zip') as tmp:
            file.save(tmp.name)
            tmp_path = tmp.name

        restored_files = []

        with zipfile.ZipFile(tmp_path, 'r') as zf:
            all_files = zf.namelist()
            print(f"Restore: zip contains files: {all_files}")
            print(f"Restore: DATA_DIR = {DATA_DIR}")

            for name in all_files:
                # Only allow specific JSON files
                if name in ['schedule.json', 'schedule_spring_2027.json', 'course_history.json']:
                    content = zf.read(name)
                    # Validate it's valid JSON
                    json.loads(content)

                    target_path = os.path.join(DATA_DIR, name)
                    print(f"Restore: writing {name} to {target_path}")
                    os.makedirs(DATA_DIR, exist_ok=True)
                    with open(target_path, 'wb') as f:
                        f.write(content)
                    restored_files.append(name)
                    print(f"Restore: successfully wrote {name}")

        # Clean up temp file
        os.unlink(tmp_path)

        # Broadcast update to all clients
        socketio.emit('data_restored', {
            'files': restored_files,
            'timestamp': datetime.now().isoformat()
        })

        return jsonify({'success': True, 'restored': restored_files})

    except json.JSONDecodeError:
        return jsonify({'success': False, 'error': 'Invalid JSON in zip file'}), 400
    except zipfile.BadZipFile:
        return jsonify({'success': False, 'error': 'Invalid zip file'}), 400
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/ai-recommendations', methods=['POST'])
def get_ai_recommendations():
    """Get AI-powered schedule recommendations using OpenAI."""
    openai_key = os.environ.get('OPENAI_API_KEY')
    if not openai_key:
        return jsonify({'success': False, 'error': 'OpenAI API key not configured'}), 500

    data = request.json or {}
    term = data.get('term', DEFAULT_TERM)
    schedule = load_schedule(term)
    course_history = load_course_history()

    # Build faculty schedule summary
    faculty_schedules = {}
    for course in schedule['courses']:
        instructor = course.get('instructor', '')
        if instructor and instructor not in ['', 'Faculty', 'TBA']:
            if instructor not in faculty_schedules:
                faculty_schedules[instructor] = []
            faculty_schedules[instructor].append({
                'course': f"{course['code']} {course['number']}-{course['section']}",
                'name': course.get('name', ''),
                'slot': course.get('slotId', 'Unscheduled'),
                'days': course.get('days', ''),
                'time': f"{course.get('startTime', '')} - {course.get('endTime', '')}"
            })

    # Build slot occupancy summary
    slot_occupancy = {}
    for course in schedule['courses']:
        slot = course.get('slotId')
        if slot:
            if slot not in slot_occupancy:
                slot_occupancy[slot] = []
            slot_occupancy[slot].append(f"{course['code']} {course['number']} ({course.get('instructor', 'TBA')})")

    # Get prerequisite information from course history
    prereq_info = {}
    for course_key, course_data in course_history.items():
        desc = course_data.get('description', '')
        if 'prerequisite' in desc.lower() or 'prereq' in desc.lower():
            prereq_info[course_key] = desc[:500]  # Limit length

    term_label = "Fall 2026" if term == 'fall-2026' else "Spring 2027"

    prompt = f"""You are an expert academic schedule advisor for the Delaplaine School of Business at Hood College. Analyze the following {term_label} course schedule and provide specific, actionable recommendations.

## SCHEDULE DATA

### All Scheduled Courses:
{json.dumps([{
    'id': c['id'],
    'course': f"{c['code']} {c['number']}-{c['section']}",
    'name': c.get('name', ''),
    'instructor': c.get('instructor', 'TBA'),
    'slot': c.get('slotId', 'Unscheduled'),
    'room': c.get('room', '')
} for c in schedule['courses']], indent=2)}

### Faculty Teaching Loads:
{json.dumps(faculty_schedules, indent=2)}

### Time Slot Occupancy:
{json.dumps(slot_occupancy, indent=2)}

### Course Prerequisite Information (from catalog):
{json.dumps(prereq_info, indent=2) if prereq_info else "No prerequisite data available"}

## TIME SLOT REFERENCE:
- MW-A: Mon/Wed 8:15-9:40 AM
- MW-B: Mon/Wed 9:50-11:15 AM
- MW-C: Mon/Wed 11:30 AM-12:55 PM
- MW-D: Mon/Wed 1:05-2:30 PM
- MW-E: Mon/Wed 2:40-4:05 PM
- TR-G: Tue/Thu 8:15-9:40 AM
- TR-H: Tue/Thu 9:50-11:15 AM
- TR-I: Tue/Thu 11:25 AM-12:50 PM
- TR-J: Tue/Thu 2:00-3:25 PM
- TR-K: Tue/Thu 3:35-5:00 PM
- M-EVE: Monday 6:15-9:00 PM
- T-EVE: Tuesday 6:15-9:00 PM
- W-EVE: Wednesday 6:15-9:00 PM
- TR-EVE: Thursday 6:15-9:00 PM
- SAT: Saturday
- ASYNCH: Online Asynchronous

## ANALYSIS CRITERIA:

1. **Faculty Workload & Scheduling**:
   - Look for faculty teaching back-to-back classes (no break between)
   - Identify faculty with classes on both MW and TR in the same time period
   - Flag any faculty with 4+ courses
   - Note if faculty are teaching at inconvenient times (e.g., morning + evening same day)

2. **Prerequisite Chains**:
   - MGMT 205 (Principles of Management) is a prereq for many 300+ level MGMT courses
   - MGMT 281 (Financial Accounting) is a prereq for MGMT 284 (Managerial Accounting)
   - MGMT 284 is a prereq for MGMT 402 (Business Finance)
   - ECON 205/206 are prereqs for upper-level ECON courses
   - Courses in a prereq chain should NOT be scheduled at the same time

3. **Student Schedule Patterns**:
   - Core courses (MGMT 205, 281, 284, ECON 205, 206) should have multiple sections at different times
   - Upper-level courses in the same major should not conflict
   - Consider that students often take related courses together (e.g., MGMT 301 + MGMT 306)

4. **Room & Slot Conflicts**:
   - Identify any slots with too many courses (potential room shortage)
   - Note empty or underutilized time slots

5. **Evening/Weekend Courses**:
   - Graduate courses (5xx, 6xx) typically meet evenings
   - Ensure graduate courses don't conflict with each other

## OUTPUT FORMAT:

Provide your response in this exact structure:

### CRITICAL ISSUES (Must Address)
[List any serious conflicts or problems that need immediate attention]

### FACULTY WORKLOAD CONCERNS
[Analyze each faculty member's schedule for issues]

### PREREQUISITE CONFLICTS
[Identify courses scheduled at the same time where one is a prerequisite for the other]

### STUDENT SCHEDULE RECOMMENDATIONS
[Suggestions to improve student access to courses]

### SLOT OPTIMIZATION
[Recommendations for moving specific courses to better time slots]

### SUMMARY
[3-5 bullet points with the most important changes to make]

Be specific - mention exact course IDs, faculty names, and time slots in your recommendations."""

    try:
        response = requests.post(
            'https://api.openai.com/v1/chat/completions',
            headers={
                'Authorization': f'Bearer {openai_key}',
                'Content-Type': 'application/json'
            },
            json={
                'model': 'gpt-4o',
                'messages': [
                    {'role': 'system', 'content': 'You are an expert academic schedule advisor. Provide clear, actionable recommendations.'},
                    {'role': 'user', 'content': prompt}
                ],
                'temperature': 0.7,
                'max_tokens': 4000
            },
            timeout=60
        )

        if response.status_code != 200:
            error_data = response.json() if response.text else {}
            return jsonify({
                'success': False,
                'error': f"OpenAI API error: {error_data.get('error', {}).get('message', response.text)}"
            }), 500

        result = response.json()
        recommendations = result['choices'][0]['message']['content']

        return jsonify({
            'success': True,
            'recommendations': recommendations,
            'term': term,
            'model': 'gpt-4o'
        })

    except requests.exceptions.Timeout:
        return jsonify({'success': False, 'error': 'Request timed out'}), 504
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@socketio.on('connect')
def handle_connect():
    """Handle new WebSocket connection."""
    print(f'Client connected: {request.sid}')
    emit('connected', {'status': 'connected', 'sid': request.sid})


@socketio.on('disconnect')
def handle_disconnect():
    """Handle WebSocket disconnection."""
    print(f'Client disconnected: {request.sid}')


@socketio.on('move_course')
def handle_move_course(data):
    """Handle course move from client."""
    term = data.get('term', DEFAULT_TERM)
    schedule = load_schedule(term)

    course_id = data.get('courseId')
    new_slot_id = data.get('slotId')

    for course in schedule['courses']:
        if course['id'] == course_id:
            course['slotId'] = new_slot_id
            break

    save_schedule(schedule, term)

    emit('schedule_update', {
        'courseId': course_id,
        'slotId': new_slot_id,
        'term': term,
        'timestamp': datetime.now().isoformat()
    }, broadcast=True)


@socketio.on('request_sync')
def handle_sync_request(data=None):
    """Send full schedule to requesting client."""
    term = DEFAULT_TERM
    if data and isinstance(data, dict):
        term = data.get('term', DEFAULT_TERM)
    emit('full_sync', load_schedule(term))


if __name__ == '__main__':
    socketio.run(app, debug=True, host='0.0.0.0', port=5001, use_reloader=False, allow_unsafe_werkzeug=True)
